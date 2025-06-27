"""
Excel export module for regulatory document processing results.
"""

import os
import logging
from datetime import datetime
from typing import List, Dict, Any, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Handles export of processed documents to Excel format."""
    
    def __init__(self, max_cell_length: int = 32767):
        """
        Initialize the Excel exporter.
        
        Args:
            max_cell_length: Maximum length for Excel cell content
        """
        self.max_cell_length = max_cell_length
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export_to_excel(self, documents: List[Dict[str, Any]], output_path: str, 
                       include_full_text: bool = False, include_statistics: bool = True,
                       include_validation: bool = True, include_articles: bool = True,
                       user_info: Optional[Dict[str, str]] = None):
        """
        Export processed documents to Excel file.
        
        Args:
            documents: List of processed document dictionaries
            output_path: Path for output Excel file
            include_full_text: Whether to include full text in export
            include_statistics: Whether to include statistics sheet
            include_validation: Whether to include validation results
            include_articles: Whether to include article breakdown
            user_info: User information to include in metadata
        """
        try:
            # Create output directory if it doesn't exist
            output_dir = os.path.dirname(output_path)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)
                logger.info(f"Created output directory: {output_dir}")
            
            logger.info(f"Starting Excel export to: {output_path}")
            logger.info(f"Export options - Articles: {include_articles}, Validation: {include_validation}, Full text: {include_full_text}")
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                logger.info("Exporting document metadata...")
                self._export_metadata(documents, writer, user_info)
                
                # Export improved data sheet with restructured content
                logger.info("Exporting restructured data...")
                self._export_restructured_data(documents, writer)
                
                if include_validation:
                    logger.info("Exporting validation scores...")
                    self._export_validation_scores(documents, writer)
                
                if include_statistics:
                    logger.info("Exporting statistics...")
                    self._export_statistics(documents, writer)
                
                logger.info("Exporting summary...")
                self._export_summary(documents, writer, user_info)
                
                logger.info("Formatting workbook...")
                self._format_workbook(writer.book)
            
            # Verify file was created
            if os.path.exists(output_path):
                file_size = os.path.getsize(output_path)
                logger.info(f"✅ Successfully exported {len(documents)} documents to {output_path} (Size: {file_size} bytes)")
            else:
                logger.error(f"❌ File was not created at {output_path}")
                raise FileNotFoundError(f"Export file was not created: {output_path}")
            
        except Exception as e:
            logger.error(f"Failed to export to Excel: {e}")
            logger.error(f"Output path: {output_path}")
            logger.error(f"Output directory exists: {os.path.exists(os.path.dirname(output_path)) if os.path.dirname(output_path) else 'No directory'}")
            raise
    
    def _export_metadata(self, documents: List[Dict[str, Any]], writer, user_info: Optional[Dict[str, str]] = None):
        """Export document metadata to Excel sheet."""
        metadata_rows = []
        
        for doc in documents:
            metadata = doc.get('metadata', {}).copy()
            stats = doc.get('statistics', {})
            
            row = {
                'File Name': metadata.get('file_name', ''),
                'Document Type': metadata.get('document_type', ''),
                'File Path': metadata.get('file_path', ''),
                'File Size': metadata.get('file_size', 0),
                'Page Count': metadata.get('page_count', 0),
                'Title': metadata.get('title', ''),
                'Author': metadata.get('author', ''),
                'Creation Date': metadata.get('creation_date', ''),
                'Extraction Date': metadata.get('extraction_date', ''),
                'Total Characters': stats.get('total_characters', 0),
                'Total Words': stats.get('total_words', 0),
                'Total Chunks': stats.get('total_chunks', 0),
                'Extraction Success': stats.get('extraction_success', False),
                'File Hash': metadata.get('file_hash', '')
            }
            
            # Add user information if provided
            if user_info:
                row['Processed By'] = f"{user_info.get('name', '')} {user_info.get('surname', '')}".strip()
                row['Processing Date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            metadata_rows.append(row)
        
        df = pd.DataFrame(metadata_rows)
        df.to_excel(writer, sheet_name='Document_Metadata', index=False)
    
    def _export_restructured_data(self, documents: List[Dict[str, Any]], writer):
        """Export restructured and enhanced document data."""
        all_data = []
        
        for doc in documents:
            # Extract comprehensive article information
            articles = self._extract_comprehensive_articles(doc)
            
            # Create restructured document dictionary
            doc_dict = self._create_document_dictionary(doc, articles)
            
            # Enhance with document structure if available
            if 'document_structure' in doc:
                doc_dict = self._enhance_with_document_structure(doc_dict, doc['document_structure'])
            
            # Apply AI enhancement if available
            if hasattr(self, 'ai_enhancer'):
                doc_dict = self._enhance_with_ai(doc_dict)
            
            # Flatten for Excel export
            flattened_data = self._flatten_document_structure(doc_dict)
            all_data.extend(flattened_data)
        
        if all_data:
            df = pd.DataFrame(all_data)
            # Reorder columns for better readability
            priority_cols = ['Document', 'Type', 'Article_Number', 'Article_Title', 
                           'Article_Content', 'Materiality', 'Document_Structure',
                           'Regulatory_Complexity', 'Key_Obligations', 
                           'Compliance_Requirements', 'Penalties', 'Key_References']
            other_cols = [col for col in df.columns if col not in priority_cols]
            ordered_cols = [col for col in priority_cols if col in df.columns] + other_cols
            df = df[ordered_cols]
            
            df.to_excel(writer, sheet_name='Data', index=False)
    
    def _extract_comprehensive_articles(self, doc: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Extract all articles comprehensively from document."""
        articles = []
        text = doc.get('cleaned_text', '') or doc.get('full_text', '')
        
        if not text:
            return articles
        
        # Enhanced regex patterns for article extraction
        article_patterns = [
            r'Article\s+(\d+(?:\.\d+)?(?:\s*[a-z])?(?:\s*bis|\s*ter|\s*quater)?)[^\n]*\n((?:(?!Article\s+\d+)[\s\S])*)',
            r'ARTICLE\s+(\d+(?:\.\d+)?(?:\s*[a-z])?(?:\s*bis|\s*ter|\s*quater)?)[^\n]*\n((?:(?!ARTICLE\s+\d+)[\s\S])*)',
            r'Art\.\s*(\d+(?:\.\d+)?(?:\s*[a-z])?(?:\s*bis|\s*ter|\s*quater)?)[^\n]*\n((?:(?!Art\.\s*\d+)[\s\S])*)',
            r'(\d+)\.\s*[-–]\s*((?:(?!\d+\.\s*[-–])[\s\S])*)',
            r'(\d+)\)\s*((?:(?!\d+\))[\s\S])*)'
        ]
        
        import re
        
        for pattern in article_patterns:
            matches = re.finditer(pattern, text, re.MULTILINE | re.IGNORECASE)
            for match in matches:
                article_num = match.group(1).strip()
                article_content = match.group(2).strip()
                
                # Clean and normalize content
                article_content = re.sub(r'\s+', ' ', article_content)
                article_content = article_content[:5000]  # Limit length
                
                if article_content and len(article_content) > 20:  # Minimum content length
                    articles.append({
                        'number': article_num,
                        'content': article_content,
                        'source_pattern': pattern,
                        'materiality': self._assess_materiality(article_content)
                    })
        
        # Remove duplicates based on article number
        seen = set()
        unique_articles = []
        for article in articles:
            if article['number'] not in seen:
                seen.add(article['number'])
                unique_articles.append(article)
        
        # Sort by article number
        unique_articles.sort(key=lambda x: self._parse_article_number_for_sort(x['number']))
        
        return unique_articles
    
    def _parse_article_number_for_sort(self, article_number: str) -> tuple:
        """Parse article number for sorting."""
        import re
        
        # Remove 'Article' prefix if present
        number_str = re.sub(r'^(Article|ARTICLE|Art\.)\s*', '', article_number, flags=re.IGNORECASE).strip()
        
        # Handle various formats
        match = re.match(r'(\d+)(?:\.(\d+))?(?:\s*([a-z]))?(?:\s*(bis|ter|quater))?', number_str, re.IGNORECASE)
        if match:
            main_num = int(match.group(1))
            sub_num = int(match.group(2)) if match.group(2) else 0
            letter = ord(match.group(3).lower()) - ord('a') if match.group(3) else 0
            suffix = match.group(4)
            suffix_val = 0
            if suffix:
                suffix_map = {'bis': 1, 'ter': 2, 'quater': 3}
                suffix_val = suffix_map.get(suffix.lower(), 0)
            return (main_num, sub_num, letter, suffix_val)
        
        return (9999, 0, 0, 0)
    
    def _assess_materiality(self, content: str) -> str:
        """Assess the materiality of an article based on keywords."""
        high_priority_keywords = [
            'obligation', 'interdit', 'doit', 'sanction', 'pénalité', 'amende',
            'suspension', 'retrait', 'révocation', 'capital', 'fonds propres',
            'ratio', 'limite', 'maximum', 'minimum', 'délai', 'immédiat'
        ]
        
        medium_priority_keywords = [
            'peut', 'devrait', 'recommandé', 'rapport', 'information',
            'communication', 'notification', 'déclaration', 'procédure'
        ]
        
        content_lower = content.lower()
        
        high_count = sum(1 for keyword in high_priority_keywords if keyword in content_lower)
        medium_count = sum(1 for keyword in medium_priority_keywords if keyword in content_lower)
        
        if high_count >= 2:
            return 'HIGH'
        elif high_count >= 1 or medium_count >= 2:
            return 'MEDIUM'
        else:
            return 'LOW'
    
    def _create_document_dictionary(self, doc: Dict[str, Any], articles: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Create a structured dictionary representation of the document."""
        metadata = doc.get('metadata', {})
        
        return {
            'document_info': {
                'name': metadata.get('file_name', ''),
                'type': metadata.get('document_type', ''),
                'date': metadata.get('creation_date', ''),
                'pages': metadata.get('page_count', 0)
            },
            'articles': articles,
            'summary': {
                'total_articles': len(articles),
                'high_priority_articles': sum(1 for a in articles if a.get('materiality') == 'HIGH'),
                'medium_priority_articles': sum(1 for a in articles if a.get('materiality') == 'MEDIUM'),
                'low_priority_articles': sum(1 for a in articles if a.get('materiality') == 'LOW')
            },
            'compliance_requirements': self._extract_compliance_requirements(doc, articles),
            'key_obligations': self._extract_key_obligations(articles),
            'penalties': self._extract_penalties(articles)
        }
    
    def _extract_compliance_requirements(self, doc: Dict[str, Any], articles: List[Dict[str, Any]]) -> List[str]:
        """Extract compliance requirements from articles."""
        requirements = []
        compliance_keywords = ['doit', 'obligation', 'exigence', 'requis', 'nécessaire']
        
        for article in articles:
            content = article['content'].lower()
            for keyword in compliance_keywords:
                if keyword in content:
                    # Extract sentence containing the keyword
                    sentences = article['content'].split('.')
                    for sentence in sentences:
                        if keyword in sentence.lower():
                            requirements.append(f"Art. {article['number']}: {sentence.strip()}")
                            break
                    break
        
        return requirements[:20]  # Limit to top 20 requirements
    
    def _extract_key_obligations(self, articles: List[Dict[str, Any]]) -> List[str]:
        """Extract key obligations from articles."""
        obligations = []
        obligation_patterns = [
            r'(?:doit|doivent|est tenu|sont tenus)\s+(?:de\s+)?([^.]+)',
            r'(?:obligation|obligatoire)\s+(?:de\s+)?([^.]+)',
            r'il est (?:interdit|prohibé)\s+(?:de\s+)?([^.]+)'
        ]
        
        import re
        
        for article in articles:
            for pattern in obligation_patterns:
                matches = re.finditer(pattern, article['content'], re.IGNORECASE)
                for match in matches:
                    obligation = match.group(0).strip()
                    if len(obligation) > 20:  # Minimum length
                        obligations.append(f"Art. {article['number']}: {obligation}")
        
        return obligations[:15]  # Limit to top 15 obligations
    
    def _extract_penalties(self, articles: List[Dict[str, Any]]) -> List[str]:
        """Extract penalties and sanctions from articles."""
        penalties = []
        penalty_keywords = ['sanction', 'pénalité', 'amende', 'suspension', 'retrait', 'révocation']
        
        for article in articles:
            content = article['content'].lower()
            for keyword in penalty_keywords:
                if keyword in content:
                    # Extract sentence containing the keyword
                    sentences = article['content'].split('.')
                    for sentence in sentences:
                        if keyword in sentence.lower():
                            penalties.append(f"Art. {article['number']}: {sentence.strip()}")
                            break
                    break
        
        return penalties[:10]  # Limit to top 10 penalties
    
    def _enhance_with_document_structure(self, doc_dict: Dict[str, Any], structure: Dict[str, Any]) -> Dict[str, Any]:
        """Enhance document dictionary with extracted structure information."""
        # Add document structure information
        doc_dict['structure_info'] = {
            'document_type_identified': structure.get('document_type', 'Unknown'),
            'sections_found': len(structure.get('sections', [])),
            'hierarchy_levels': {
                'titres': len(structure.get('hierarchy', {}).get('titres', [])),
                'chapitres': len(structure.get('hierarchy', {}).get('chapitres', [])),
                'sections': len(structure.get('hierarchy', {}).get('sections', []))
            },
            'regulatory_references': len(structure.get('references', [])),
            'extracted_metadata': structure.get('metadata_extracted', {})
        }
        
        # Enhance article context with structural information
        if 'articles' in doc_dict:
            for article in doc_dict['articles']:
                # Find which section/chapter this article belongs to
                article_context = self._find_article_context(article, structure)
                if article_context:
                    article['structural_context'] = article_context
        
        # Add regulatory compliance insights
        doc_dict['compliance_insights'] = self._extract_compliance_insights(structure)
        
        return doc_dict
    
    def _find_article_context(self, article: Dict[str, Any], structure: Dict[str, Any]) -> Optional[Dict[str, str]]:
        """Find the structural context (TITRE, CHAPITRE, Section) for an article."""
        # This would need the article position to determine context
        # For now, return basic structure information
        hierarchy = structure.get('hierarchy', {})
        
        context = {}
        if hierarchy.get('titres'):
            context['titre'] = hierarchy['titres'][0].get('title', 'Unknown')
        if hierarchy.get('chapitres'):
            context['chapitre'] = hierarchy['chapitres'][0].get('title', 'Unknown')
        if hierarchy.get('sections'):
            context['section'] = hierarchy['sections'][0].get('title', 'Unknown')
        
        return context if context else None
    
    def _extract_compliance_insights(self, structure: Dict[str, Any]) -> Dict[str, Any]:
        """Extract compliance-related insights from document structure."""
        insights = {
            'regulatory_complexity': 'Low',
            'key_references': [],
            'document_completeness': 'Complete'
        }
        
        # Assess regulatory complexity based on structure
        total_sections = len(structure.get('sections', []))
        total_hierarchy = sum(len(h) for h in structure.get('hierarchy', {}).values())
        
        if total_sections + total_hierarchy > 10:
            insights['regulatory_complexity'] = 'High'
        elif total_sections + total_hierarchy > 5:
            insights['regulatory_complexity'] = 'Medium'
        
        # Extract key regulatory references
        references = structure.get('references', [])
        insights['key_references'] = [
            f"{ref['type']}: {ref['reference']}" 
            for ref in references[:5]  # Top 5 references
        ]
        
        # Assess document completeness
        sections = structure.get('sections', [])
        expected_sections = ['Header', 'Preamble', 'Decide', 'Articles', 'Signature']
        found_sections = [s['name'] for s in sections]
        missing_sections = [s for s in expected_sections if s not in found_sections]
        
        if missing_sections:
            insights['document_completeness'] = f"Missing: {', '.join(missing_sections)}"
        
        return insights
    
    def _enhance_with_ai(self, doc_dict: Dict[str, Any]) -> Dict[str, Any]:
        """Enhance document dictionary with AI if available."""
        # This method would use AI to enhance the extraction
        # For now, return as-is
        return doc_dict
    
    def _flatten_document_structure(self, doc_dict: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Flatten the document structure for Excel export."""
        flattened = []
        doc_info = doc_dict['document_info']
        
        # Add summary row
        structure_info = doc_dict.get('structure_info', {})
        compliance_insights = doc_dict.get('compliance_insights', {})
        
        summary_row = {
            'Document': doc_info['name'],
            'Type': doc_info['type'],
            'Article_Number': 'SUMMARY',
            'Article_Title': 'Document Summary',
            'Article_Content': f"Total Articles: {doc_dict['summary']['total_articles']}, "
                             f"High Priority: {doc_dict['summary']['high_priority_articles']}, "
                             f"Medium Priority: {doc_dict['summary']['medium_priority_articles']}, "
                             f"Low Priority: {doc_dict['summary']['low_priority_articles']}",
            'Materiality': 'INFO',
            'Key_Obligations': '; '.join(doc_dict['key_obligations'][:3]) if doc_dict['key_obligations'] else '',
            'Compliance_Requirements': '; '.join(doc_dict['compliance_requirements'][:3]) if doc_dict['compliance_requirements'] else '',
            'Penalties': '; '.join(doc_dict['penalties'][:3]) if doc_dict['penalties'] else '',
            'Document_Structure': structure_info.get('document_type_identified', 'Unknown'),
            'Regulatory_Complexity': compliance_insights.get('regulatory_complexity', 'Unknown'),
            'Document_Completeness': compliance_insights.get('document_completeness', 'Unknown'),
            'Key_References': '; '.join(compliance_insights.get('key_references', [])[:2])
        }
        flattened.append(summary_row)
        
        # Add article rows
        for article in doc_dict['articles']:
            structural_context = article.get('structural_context', {})
            context_text = []
            if structural_context.get('titre'):
                context_text.append(f"TITRE: {structural_context['titre']}")
            if structural_context.get('chapitre'):
                context_text.append(f"CHAPITRE: {structural_context['chapitre']}")
            if structural_context.get('section'):
                context_text.append(f"SECTION: {structural_context['section']}")
            
            article_row = {
                'Document': doc_info['name'],
                'Type': doc_info['type'],
                'Article_Number': f"Article {article['number']}",
                'Article_Title': f"Article {article['number']}",
                'Article_Content': self._truncate_text(article['content'], 2000),
                'Materiality': article.get('materiality', 'UNKNOWN'),
                'Key_Obligations': '',
                'Compliance_Requirements': '',
                'Penalties': '',
                'Document_Structure': '; '.join(context_text) if context_text else '',
                'Regulatory_Complexity': '',
                'Document_Completeness': '',
                'Key_References': ''
            }
            
            # Add specific requirements for this article
            for req in doc_dict['compliance_requirements']:
                if f"Art. {article['number']}:" in req:
                    article_row['Compliance_Requirements'] = req.split(':', 1)[1].strip()
                    break
            
            for obl in doc_dict['key_obligations']:
                if f"Art. {article['number']}:" in obl:
                    article_row['Key_Obligations'] = obl.split(':', 1)[1].strip()
                    break
            
            for pen in doc_dict['penalties']:
                if f"Art. {article['number']}:" in pen:
                    article_row['Penalties'] = pen.split(':', 1)[1].strip()
                    break
            
            flattened.append(article_row)
        
        return flattened
    
    def _export_chunks(self, documents: List[Dict[str, Any]], writer):
        """Export text chunks to Excel sheet."""
        all_chunks = []
        
        for doc in documents:
            chunks = doc.get('chunks', [])
            for chunk in chunks:
                chunk_row = {
                    'Chunk ID': chunk.get('chunk_id', ''),
                    'File Name': chunk.get('file_name', ''),
                    'Document Type': chunk.get('document_type', ''),
                    'Chunk Index': chunk.get('chunk_index', 0),
                    'Text': self._truncate_text(chunk.get('text', '')),
                    'Text Length': chunk.get('text_length', 0),
                    'Word Count': chunk.get('word_count', 0),
                    'Sentence Count': chunk.get('sentence_count', 0),
                    'Start Character': chunk.get('start_char', 0),
                    'Chunk Hash': chunk.get('chunk_hash', '')
                }
                
                # Add validation scores if present
                if 'validation_score' in chunk:
                    val_score = chunk['validation_score']
                    chunk_row.update({
                        'Completeness Score': val_score.get('completeness_score', 0),
                        'Reliability Score': val_score.get('reliability_score', 0),
                        'Legal Structure Score': val_score.get('legal_structure_score', 0),
                        'Overall Validation Score': val_score.get('overall_score', 0)
                    })
                
                all_chunks.append(chunk_row)
        
        df = pd.DataFrame(all_chunks)
        df.to_excel(writer, sheet_name='Text_Chunks', index=False)
    
    def _export_articles(self, documents: List[Dict[str, Any]], writer):
        """Export extracted articles to Excel sheet with specific format."""
        all_articles = []
        
        for doc in documents:
            articles = doc.get('articles', [])
            
            # Get regulation type from context or filename
            regulation_type = 'UNKNOWN'
            if hasattr(articles[0], 'context') and articles:
                regulation_type = articles[0].context.get('regulation_type', 'UNKNOWN')
            else:
                filename = doc['metadata'].get('file_name', '').upper()
                if 'COBAC' in filename:
                    regulation_type = 'Règlement COBAC'
                elif 'CEMAC' in filename:
                    regulation_type = 'Règlement CEMAC'
            
            for article in articles:
                article_row = {
                    'Regulation': regulation_type,
                    'Article X': article.number,
                    'Article Content': self._truncate_text(article.content),
                    'Materiality Assessment': article.materiality.value,
                    'Context': article.materiality_reasoning
                }
                all_articles.append(article_row)
        
        if all_articles:
            df = pd.DataFrame(all_articles)
            # Create both sheets - one with the requested name "Data"
            df.to_excel(writer, sheet_name='Data', index=False)
            # Keep the original Articles sheet for backward compatibility
            df_detailed = pd.DataFrame([{
                'Regulation': row['Regulation'],
                'Article Number': row['Article X'],
                'Content': row['Article Content'],
                'Materiality': row['Materiality Assessment'],
                'Materiality Reasoning': row['Context'],
                'Word Count': len(row['Article Content'].split()),
                'Character Count': len(row['Article Content'])
            } for row in all_articles])
            df_detailed.to_excel(writer, sheet_name='Articles', index=False)
    
    def _export_validation_scores(self, documents: List[Dict[str, Any]], writer):
        """Export validation scores and results."""
        validation_rows = []
        
        for doc in documents:
            val_results = doc.get('validation_results', {})
            doc_val = val_results.get('document_validation', {})
            
            if doc_val:
                val_row = {
                    'File Name': doc['metadata'].get('file_name', ''),
                    'Document Type': doc['metadata'].get('document_type', ''),
                    'Overall Document Score': doc_val.get('overall_score', 0),
                    'Extraction Quality Score': doc_val.get('extraction_quality', {}).get('score', 0),
                    'Chunk Boundary Integrity': doc_val.get('chunk_boundary_integrity', {}).get('average_score', 0),
                    'Legal Structure Score': doc_val.get('legal_structure', {}).get('score', 0),
                    'Completeness Score': doc_val.get('completeness', {}).get('score', 0),
                    'Total Articles': doc_val.get('legal_structure', {}).get('total_articles', 0),
                    'Validation Timestamp': val_results.get('validation_timestamp', '')
                }
                
                # Add extraction quality issues
                issues = doc_val.get('extraction_quality', {}).get('issues', [])
                val_row['Extraction Issues'] = '; '.join(issues) if issues else 'None'
                
                # Add legal structure issues
                struct_issues = doc_val.get('legal_structure', {}).get('issues', [])
                val_row['Structure Issues'] = '; '.join(struct_issues) if struct_issues else 'None'
                
                validation_rows.append(val_row)
        
        if validation_rows:
            df = pd.DataFrame(validation_rows)
            df.to_excel(writer, sheet_name='Validation_Results', index=False)
    
    def _export_full_texts(self, documents: List[Dict[str, Any]], writer):
        """Export full document texts to Excel sheet."""
        text_rows = []
        
        for doc in documents:
            text_row = {
                'File Name': doc['metadata'].get('file_name', ''),
                'Document Type': doc['metadata'].get('document_type', ''),
                'Page Count': doc['metadata'].get('page_count', 0),
                'Cleaned Text': self._truncate_text(doc.get('cleaned_text', '')),
                'Text Length': len(doc.get('cleaned_text', '')),
                'Word Count': doc['statistics'].get('total_words', 0)
            }
            text_rows.append(text_row)
        
        df = pd.DataFrame(text_rows)
        df.to_excel(writer, sheet_name='Full_Texts', index=False)
    
    def _export_statistics(self, documents: List[Dict[str, Any]], writer):
        """Export processing statistics to Excel sheet."""
        doc_types = {}
        total_pages = 0
        total_chunks = 0
        successful_extractions = 0
        
        for doc in documents:
            doc_type = doc['metadata'].get('document_type', 'OTHER')
            doc_types[doc_type] = doc_types.get(doc_type, 0) + 1
            total_pages += doc['metadata'].get('page_count', 0)
            total_chunks += doc['statistics'].get('total_chunks', 0)
            if doc['statistics'].get('extraction_success', False):
                successful_extractions += 1
        
        stats_data = []
        
        stats_data.append({
            'Metric': 'Total Documents Processed',
            'Value': len(documents)
        })
        
        stats_data.append({
            'Metric': 'Successful Extractions',
            'Value': successful_extractions
        })
        
        stats_data.append({
            'Metric': 'Failed Extractions',
            'Value': len(documents) - successful_extractions
        })
        
        stats_data.append({
            'Metric': 'Total Pages',
            'Value': total_pages
        })
        
        stats_data.append({
            'Metric': 'Total Chunks Generated',
            'Value': total_chunks
        })
        
        stats_data.append({
            'Metric': 'Average Pages per Document',
            'Value': round(total_pages / len(documents), 2) if documents else 0
        })
        
        stats_data.append({
            'Metric': 'Average Chunks per Document',
            'Value': round(total_chunks / len(documents), 2) if documents else 0
        })
        
        for doc_type, count in sorted(doc_types.items()):
            stats_data.append({
                'Metric': f'Documents of Type: {doc_type}',
                'Value': count
            })
        
        df = pd.DataFrame(stats_data)
        df.to_excel(writer, sheet_name='Statistics', index=False)
    
    def _export_summary(self, documents: List[Dict[str, Any]], writer, user_info: Optional[Dict[str, str]] = None):
        """Export processing summary to Excel sheet."""
        summary_data = {
            'Processing Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'Total Documents': len(documents),
            'Successful Extractions': sum(1 for d in documents if d['statistics'].get('extraction_success', False)),
            'Total Pages': sum(d['metadata'].get('page_count', 0) for d in documents),
            'Total Chunks': sum(d['statistics'].get('total_chunks', 0) for d in documents),
            'Total Characters': sum(d['statistics'].get('total_characters', 0) for d in documents),
            'Total Words': sum(d['statistics'].get('total_words', 0) for d in documents),
        }
        
        # Add user information if provided
        if user_info:
            summary_data['Processed By'] = f"{user_info.get('name', '')} {user_info.get('surname', '')}".strip()
            summary_data['User First Name'] = user_info.get('name', '')
            summary_data['User Last Name'] = user_info.get('surname', '')
        
        df = pd.DataFrame([summary_data])
        df.to_excel(writer, sheet_name='Summary', index=False)
    
    def _format_workbook(self, workbook):
        """Apply formatting to the workbook."""
        # Format specific sheets with enhanced styling
        for sheet_name in ['Data', 'Document_Metadata', 'Summary']:
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                self._format_enhanced_sheet(sheet)
        
        # Format other sheets normally
        for sheet in workbook.worksheets:
            if sheet.title not in ['Data', 'Document_Metadata', 'Summary']:
                self._format_sheet(sheet)
    
    def _format_enhanced_sheet(self, sheet):
        """Apply enhanced formatting to important sheets."""
        # Header formatting
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = PatternFill(start_color="1A8FD1", end_color="1A8FD1", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self.border
        
        # Apply zebra striping for better readability
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            if row_idx % 2 == 0:
                for cell in row:
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            
            # Apply borders to all cells
            for cell in row:
                cell.border = self.border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        # Auto-adjust column widths with constraints
        self._auto_adjust_columns(sheet)
        
        # Freeze the header row
        sheet.freeze_panes = 'A2'
    
    def _format_sheet(self, sheet):
        """Apply formatting to a sheet."""
        for cell in sheet[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                    cell.border = self.border
                    if cell.row > 1:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
        sheet.row_dimensions[1].height = 30
    
    def _truncate_text(self, text: str, max_length: Optional[int] = None) -> str:
        """Truncate text to fit Excel cell limit."""
        max_len = max_length or self.max_cell_length
        if len(text) > max_len:
            return text[:max_len - 3] + "..."
        return text
    
    def _auto_adjust_columns(self, sheet):
        """Auto-adjust column widths with constraints."""
        for column in sheet.columns:
            max_length = 0
            column_letter = None
            
            for cell in column:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                    if not column_letter:
                        column_letter = cell.column_letter
            
            if column_letter:
                # Set reasonable constraints
                if 'Content' in str(sheet[f'{column_letter}1'].value):
                    adjusted_width = min(max_length * 0.8, 80)  # Wider for content
                elif 'Number' in str(sheet[f'{column_letter}1'].value):
                    adjusted_width = min(max_length + 2, 15)  # Narrower for numbers
                else:
                    adjusted_width = min(max_length + 2, 40)  # Standard width
                
                sheet.column_dimensions[column_letter].width = adjusted_width


class CSVExporter:
    """Alternative exporter for CSV format."""
    
    def export_chunks_to_csv(self, documents: List[Dict[str, Any]], output_path: str):
        """Export chunks to CSV file."""
        all_chunks = []
        
        for doc in documents:
            for chunk in doc.get('chunks', []):
                chunk_data = {
                    'chunk_id': chunk.get('chunk_id', ''),
                    'file_name': chunk.get('file_name', ''),
                    'document_type': chunk.get('document_type', ''),
                    'chunk_index': chunk.get('chunk_index', 0),
                    'text': chunk.get('text', '').replace('\n', ' '),
                    'text_length': chunk.get('text_length', 0),
                    'word_count': chunk.get('word_count', 0)
                }
                all_chunks.append(chunk_data)
        
        df = pd.DataFrame(all_chunks)
        df.to_csv(output_path, index=False, encoding='utf-8')
        logger.info(f"Exported {len(all_chunks)} chunks to CSV: {output_path}")