"""
Client-oriented Excel exporter for regulatory documents.
Produces business-friendly output focused on compliance priorities.
"""

import os
from datetime import datetime
from typing import List, Dict, Any, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import logging

logger = logging.getLogger(__name__)


class ClientExcelExporter:
    """Export regulatory data in a client-friendly Excel format."""
    
    def __init__(self, max_cell_length: int = 32767):
        self.materiality_colors = {
            'CRITICAL': 'FF0000',  # Red
            'HIGH': 'FF6600',      # Orange
            'MEDIUM': 'FFCC00',    # Yellow
            'LOW': '00CC00'        # Green
        }
        
        self.materiality_descriptions = {
            'CRITICAL': 'Immediate action required - Core compliance requirement',
            'HIGH': 'High priority - Significant compliance impact',
            'MEDIUM': 'Medium priority - Standard compliance requirement',
            'LOW': 'Low priority - Best practice recommendation'
        }
    
    def export_client_report(
        self, 
        documents: List[Dict[str, Any]], 
        output_path: str,
        user_info: Optional[Dict[str, str]] = None
    ):
        """Export regulatory analysis in client-friendly format."""
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create sheets
        self._create_executive_summary(wb, documents, user_info)
        self._create_document_overview(wb, documents)
        self._create_compliance_articles(wb, documents)
        self._create_priority_actions(wb, documents)
        self._create_glossary(wb)
        
        # Save workbook
        wb.save(output_path)
        logger.info(f"Client report exported to: {output_path}")
    
    def _create_executive_summary(self, wb: Workbook, documents: List[Dict[str, Any]], user_info: Optional[Dict[str, str]]):
        """Create executive summary sheet."""
        ws = wb.create_sheet("Executive Summary", 0)
        
        # Title
        ws['A1'] = "Regulatory Compliance Analysis"
        ws['A1'].font = Font(size=18, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill("solid", fgColor="1F4E79")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells('A1:E1')
        ws.row_dimensions[1].height = 40
        
        # Report info
        row = 3
        ws[f'A{row}'] = "Report Generated:"
        ws[f'B{row}'] = datetime.now().strftime("%B %d, %Y")
        
        if user_info:
            row += 1
            ws[f'A{row}'] = "Prepared for:"
            ws[f'B{row}'] = f"{user_info.get('first_name', '')} {user_info.get('last_name', '')}"
        
        # Summary statistics
        row += 2
        ws[f'A{row}'] = "SUMMARY STATISTICS"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws.merge_cells(f'A{row}:C{row}')
        
        # Calculate statistics
        total_docs = len(documents)
        total_articles = sum(len(doc.get('articles', [])) for doc in documents)
        
        # Count by materiality
        materiality_counts = {'CRITICAL': 0, 'HIGH': 0, 'MEDIUM': 0, 'LOW': 0}
        for doc in documents:
            for article in doc.get('articles', []):
                level = article.get('materiality', 'MEDIUM')
                if level in materiality_counts:
                    materiality_counts[level] += 1
        
        # Display statistics
        stats = [
            ("Total Regulatory Documents", total_docs),
            ("Total Articles Analyzed", total_articles),
            ("Critical Priority Items", materiality_counts['CRITICAL']),
            ("High Priority Items", materiality_counts['HIGH']),
            ("Medium Priority Items", materiality_counts['MEDIUM']),
            ("Low Priority Items", materiality_counts['LOW'])
        ]
        
        row += 2
        for label, value in stats:
            ws[f'A{row}'] = label
            ws[f'C{row}'] = value
            ws[f'C{row}'].alignment = Alignment(horizontal="center")
            
            # Color code priority items
            if "Critical" in label:
                ws[f'C{row}'].fill = PatternFill("solid", fgColor="FFE5E5")
            elif "High" in label:
                ws[f'C{row}'].fill = PatternFill("solid", fgColor="FFF0E5")
            
            row += 1
        
        # Key findings section
        row += 2
        ws[f'A{row}'] = "KEY COMPLIANCE PRIORITIES"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws.merge_cells(f'A{row}:E{row}')
        
        row += 2
        priorities = [
            ("1. IMMEDIATE ACTIONS", f"{materiality_counts['CRITICAL']} critical compliance requirements need immediate attention"),
            ("2. SHORT-TERM FOCUS", f"{materiality_counts['HIGH']} high-priority items should be addressed within 30 days"),
            ("3. MEDIUM-TERM PLANNING", f"{materiality_counts['MEDIUM']} standard requirements for quarterly review"),
            ("4. BEST PRACTICES", f"{materiality_counts['LOW']} recommendations for continuous improvement")
        ]
        
        for priority, desc in priorities:
            ws[f'A{row}'] = priority
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            ws[f'B{row}'] = desc
            ws.merge_cells(f'B{row}:E{row}')
            row += 2
        
        # Format columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
    
    def _create_document_overview(self, wb: Workbook, documents: List[Dict[str, Any]]):
        """Create document overview sheet."""
        ws = wb.create_sheet("Document Overview", 1)
        
        # Headers
        headers = [
            "Document Title",
            "Regulation Type",
            "Reference Number",
            "Total Articles",
            "Critical Items",
            "High Priority",
            "Key Topics"
        ]
        
        # Style headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add document data
        row = 2
        for doc in documents:
            metadata = doc.get('metadata', {})
            articles = doc.get('articles', [])
            
            # Count priorities
            critical_count = sum(1 for a in articles if a.get('materiality') == 'CRITICAL')
            high_count = sum(1 for a in articles if a.get('materiality') == 'HIGH')
            
            # Extract key topics (first few article titles)
            topics = []
            for article in articles[:3]:
                if article.get('article_number'):
                    topics.append(article['article_number'])
            key_topics = ", ".join(topics) + ("..." if len(articles) > 3 else "")
            
            # Document info
            ws.cell(row=row, column=1, value=metadata.get('filename', 'Unknown'))
            ws.cell(row=row, column=2, value=metadata.get('document_type', 'REGULATION'))
            ws.cell(row=row, column=3, value=metadata.get('regulation_ref', 'N/A'))
            ws.cell(row=row, column=4, value=len(articles))
            ws.cell(row=row, column=5, value=critical_count)
            ws.cell(row=row, column=6, value=high_count)
            ws.cell(row=row, column=7, value=key_topics)
            
            # Highlight rows with critical items
            if critical_count > 0:
                for col in range(1, 8):
                    ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FFE5E5")
            
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add filters
        ws.auto_filter.ref = ws.dimensions
    
    def _create_compliance_articles(self, wb: Workbook, documents: List[Dict[str, Any]]):
        """Create detailed compliance articles sheet."""
        ws = wb.create_sheet("Compliance Articles", 2)
        
        # Headers
        headers = [
            "Regulation",
            "Article",
            "Article Text",
            "Compliance Priority",
            "Action Required",
            "Business Impact"
        ]
        
        # Style headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Add article data
        row = 2
        for doc in documents:
            regulation_name = doc.get('metadata', {}).get('filename', 'Unknown')
            
            for article in doc.get('articles', []):
                # Basic info
                ws.cell(row=row, column=1, value=regulation_name)
                ws.cell(row=row, column=2, value=article.get('article_number', 'N/A'))
                
                # Article text (truncated for readability)
                content = article.get('content', '')
                if len(content) > 500:
                    content = content[:497] + "..."
                ws.cell(row=row, column=3, value=content)
                ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True, vertical="top")
                
                # Priority
                materiality = article.get('materiality', 'MEDIUM')
                priority_cell = ws.cell(row=row, column=4, value=materiality)
                priority_cell.font = Font(bold=True)
                
                # Color code priority
                if materiality in self.materiality_colors:
                    priority_cell.fill = PatternFill("solid", fgColor=self.materiality_colors[materiality])
                    priority_cell.font = Font(bold=True, color="FFFFFF")
                
                # Action required
                ws.cell(row=row, column=5, value=self.materiality_descriptions.get(materiality, ''))
                ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True, vertical="top")
                
                # Business impact
                reasoning = article.get('materiality_reasoning', '')
                if len(reasoning) > 200:
                    reasoning = reasoning[:197] + "..."
                ws.cell(row=row, column=6, value=reasoning)
                ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True, vertical="top")
                
                # Row height for readability
                ws.row_dimensions[row].height = 60
                
                row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 35
        
        # Add filters
        ws.auto_filter.ref = ws.dimensions
        
        # Freeze top row
        ws.freeze_panes = 'A2'
    
    def _create_priority_actions(self, wb: Workbook, documents: List[Dict[str, Any]]):
        """Create priority actions dashboard."""
        ws = wb.create_sheet("Priority Actions", 3)
        
        # Title
        ws['A1'] = "Compliance Priority Dashboard"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Collect all articles by priority
        priority_articles = {'CRITICAL': [], 'HIGH': [], 'MEDIUM': [], 'LOW': []}
        
        for doc in documents:
            regulation_name = doc.get('metadata', {}).get('filename', 'Unknown')
            for article in doc.get('articles', []):
                materiality = article.get('materiality', 'MEDIUM')
                if materiality in priority_articles:
                    priority_articles[materiality].append({
                        'regulation': regulation_name,
                        'article': article.get('article_number', 'N/A'),
                        'summary': article.get('content', '')[:100] + "...",
                        'action': self.materiality_descriptions.get(materiality, '')
                    })
        
        # Create sections for each priority
        row = 3
        for priority in ['CRITICAL', 'HIGH', 'MEDIUM', 'LOW']:
            articles = priority_articles[priority]
            
            if articles:
                # Section header
                ws[f'A{row}'] = f"{priority} PRIORITY ({len(articles)} items)"
                ws[f'A{row}'].font = Font(size=14, bold=True, color="FFFFFF")
                ws[f'A{row}'].fill = PatternFill("solid", fgColor=self.materiality_colors[priority])
                ws.merge_cells(f'A{row}:D{row}')
                row += 1
                
                # Items
                for item in articles[:10]:  # Show top 10 per category
                    ws[f'A{row}'] = item['regulation']
                    ws[f'B{row}'] = item['article']
                    ws[f'C{row}'] = item['summary']
                    ws[f'D{row}'] = item['action']
                    
                    # Light background
                    for col in ['A', 'B', 'C', 'D']:
                        ws[f'{col}{row}'].fill = PatternFill("solid", fgColor="F5F5F5")
                    
                    row += 1
                
                row += 1  # Space between sections
        
        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 35
    
    def _create_glossary(self, wb: Workbook):
        """Create glossary sheet with compliance terms."""
        ws = wb.create_sheet("Glossary", 4)
        
        # Title
        ws['A1'] = "Regulatory Terms Glossary"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:B1')
        
        # Terms
        glossary_terms = [
            ("COBAC", "Commission Bancaire de l'Afrique Centrale - Central African Banking Commission"),
            ("CEMAC", "Communauté Économique et Monétaire de l'Afrique Centrale - Economic and Monetary Community of Central Africa"),
            ("Regulation", "Binding legislative act that must be applied in its entirety"),
            ("Instruction", "Detailed guidance on how to implement regulations"),
            ("Capital Adequacy", "Measure of a bank's capital expressed as a percentage of risk-weighted assets"),
            ("Liquidity Ratio", "Measure of a bank's ability to meet short-term obligations"),
            ("Compliance Priority", "Assessment of how critical a requirement is for regulatory compliance"),
            ("Materiality", "Significance of a requirement in terms of regulatory risk and business impact"),
            ("Risk-Weighted Assets", "Bank's assets weighted according to credit risk"),
            ("Prudential Norms", "Standards ensuring the safety and soundness of financial institutions")
        ]
        
        # Headers
        ws['A3'] = "Term"
        ws['B3'] = "Definition"
        ws['A3'].font = Font(bold=True)
        ws['B3'].font = Font(bold=True)
        
        # Add terms
        row = 4
        for term, definition in glossary_terms:
            ws[f'A{row}'] = term
            ws[f'B{row}'] = definition
            ws[f'B{row}'].alignment = Alignment(wrap_text=True)
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 80