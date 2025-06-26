"""
Base classes for exporters to reduce code duplication.
Provides common Excel formatting and styling functionality.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, Any, List
import logging

logger = logging.getLogger(__name__)


class BaseExcelExporter:
    """Base class for Excel exporters with common functionality."""
    
    def __init__(self, max_cell_length: int = 32767):
        self.max_cell_length = max_cell_length
        self.styles = self._setup_styles()
        self.colors = self._setup_colors()
    
    def _setup_styles(self) -> Dict[str, Any]:
        """Centralized style configuration."""
        return {
            'header_font': Font(bold=True, color="FFFFFF", size=12),
            'header_fill': PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
            'header_alignment': Alignment(horizontal="center", vertical="center", wrap_text=True),
            'data_font': Font(size=10),
            'data_alignment': Alignment(vertical="top", wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    
    def _setup_colors(self) -> Dict[str, str]:
        """Color scheme for different priority levels and statuses."""
        return {
            'critical': 'FF0000',    # Red
            'high': 'FF6600',        # Orange  
            'medium': 'FFCC00',      # Yellow
            'low': '00CC00',         # Green
            'success': '27ae60',     # Dark Green
            'error': 'e74c3c',       # Dark Red
            'warning': 'f39c12',     # Orange
            'info': '3498db',        # Blue
            'header': '2c3e50'       # Dark Blue
        }
    
    def _format_header_row(self, worksheet, row: int, columns: List[str]) -> None:
        """Apply consistent header formatting."""
        for col, header in enumerate(columns, 1):
            cell = worksheet.cell(row=row, column=col, value=header)
            cell.font = self.styles['header_font']
            cell.fill = self.styles['header_fill']
            cell.alignment = self.styles['header_alignment']
            cell.border = self.styles['border']
    
    def _format_data_cell(self, cell, value: Any, cell_type: str = "default") -> None:
        """Apply consistent data cell formatting."""
        # Truncate text if too long
        if isinstance(value, str) and len(value) > self.max_cell_length:
            value = value[:self.max_cell_length-3] + "..."
        
        cell.value = value
        cell.font = self.styles['data_font']
        cell.alignment = self.styles['data_alignment']
        cell.border = self.styles['border']
        
        # Apply special formatting based on cell type
        if cell_type == "priority":
            self._format_priority_cell(cell, value)
        elif cell_type == "status":
            self._format_status_cell(cell, value)
    
    def _format_priority_cell(self, cell, priority: str) -> None:
        """Format cells based on priority level."""
        priority_lower = str(priority).lower()
        
        if priority_lower == 'critical':
            cell.fill = PatternFill(start_color=self.colors['critical'], fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        elif priority_lower == 'high':
            cell.fill = PatternFill(start_color=self.colors['high'], fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        elif priority_lower == 'medium':
            cell.fill = PatternFill(start_color=self.colors['medium'], fill_type="solid")
            cell.font = Font(bold=True, color="000000")
        elif priority_lower == 'low':
            cell.fill = PatternFill(start_color=self.colors['low'], fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
    
    def _format_status_cell(self, cell, status: str) -> None:
        """Format cells based on status."""
        status_lower = str(status).lower()
        
        if 'success' in status_lower or '✓' in str(status):
            cell.fill = PatternFill(start_color=self.colors['success'], fill_type="solid")
            cell.font = Font(color="FFFFFF")
        elif 'error' in status_lower or 'failed' in status_lower or '✗' in str(status):
            cell.fill = PatternFill(start_color=self.colors['error'], fill_type="solid")
            cell.font = Font(color="FFFFFF")
        elif 'warning' in status_lower or '⚠' in str(status):
            cell.fill = PatternFill(start_color=self.colors['warning'], fill_type="solid")
            cell.font = Font(color="FFFFFF")
    
    def _auto_adjust_columns(self, worksheet, min_width: int = 10, max_width: int = 50) -> None:
        """Auto-adjust column widths for better readability."""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _add_filters_and_freeze(self, worksheet, freeze_row: int = 2) -> None:
        """Add auto-filters and freeze top rows."""
        if worksheet.max_row > 1:
            worksheet.auto_filter.ref = worksheet.dimensions
        
        if freeze_row <= worksheet.max_row:
            worksheet.freeze_panes = f'A{freeze_row}'
    
    def _create_summary_stats(self, documents: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Generate summary statistics for documents."""
        stats = {
            'total_documents': len(documents),
            'total_pages': 0,
            'total_chunks': 0,
            'total_articles': 0,
            'document_types': {},
            'priority_counts': {'CRITICAL': 0, 'HIGH': 0, 'MEDIUM': 0, 'LOW': 0},
            'processing_errors': 0
        }
        
        for doc in documents:
            metadata = doc.get('metadata', {})
            stats['total_pages'] += metadata.get('page_count', 0)
            
            chunks = doc.get('chunks', [])
            stats['total_chunks'] += len(chunks)
            
            articles = doc.get('articles', [])
            stats['total_articles'] += len(articles)
            
            # Count by document type
            doc_type = metadata.get('document_type', 'OTHER')
            stats['document_types'][doc_type] = stats['document_types'].get(doc_type, 0) + 1
            
            # Count by priority
            for article in articles:
                priority = article.get('materiality', 'MEDIUM')
                if priority in stats['priority_counts']:
                    stats['priority_counts'][priority] += 1
        
        return stats
    
    def _truncate_text(self, text: str, max_length: int = None) -> str:
        """Truncate text to specified length."""
        if max_length is None:
            max_length = self.max_cell_length
        
        if not isinstance(text, str):
            text = str(text)
        
        if len(text) <= max_length:
            return text
        
        return text[:max_length-3] + "..."
    
    def _safe_cell_value(self, value: Any) -> Any:
        """Ensure cell value is safe for Excel."""
        if value is None:
            return ""
        
        if isinstance(value, str):
            return self._truncate_text(value)
        
        if isinstance(value, (int, float)):
            return value
        
        return self._truncate_text(str(value))